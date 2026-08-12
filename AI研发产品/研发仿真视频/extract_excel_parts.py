# -*- coding: utf-8 -*-
import json
from pathlib import Path

import openpyxl

root = Path(r"C:\Users\sanye\Desktop\SMAR\AI研发产品")
p = list(root.glob("*.xlsx"))[0]
out = root / "研发仿真视频" / "out" / "_excel_parts"
out.mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(p)
ws = wb[wb.sheetnames[1]]

row_name = {}
for r in range(1, ws.max_row + 1):
    v = ws.cell(r, 3).value
    if isinstance(v, str) and v.strip():
        row_name[r] = v.strip()

mapping = []
for i, img in enumerate(ws._images):
    row = img.anchor._from.row + 1
    col = img.anchor._from.col + 1
    label = row_name.get(row) or ""
    if not label:
        for rr in range(row, max(0, row - 3), -1):
            if rr in row_name:
                label = row_name[rr]
                break
    data = img._data()
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        ext = ".png"
    elif data[:2] == b"\xff\xd8":
        ext = ".jpg"
    else:
        ext = ".png"
    safe = "".join("_" if c in '<>:"/\\|?*' else c for c in label) or f"img{i}"
    fname = f"{i:02d}_r{row}_{safe}{ext}"
    (out / fname).write_bytes(data)
    mapping.append(
        {
            "i": i,
            "row": row,
            "col": col,
            "label": label,
            "file": fname,
            "part_no": ws.cell(row, 1).value,
        }
    )

(out / "manifest.json").write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8")
(out / "rows.json").write_text(json.dumps(row_name, ensure_ascii=False, indent=2), encoding="utf-8")
print("xlsx", p.name)
print("sheet", ws.title)
print("parts", len(mapping))
for m in mapping:
    print(f"{m['i']:02d} #{m['part_no']} {m['label']} -> {m['file']}")
